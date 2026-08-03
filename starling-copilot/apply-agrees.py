#!/usr/bin/env python3
"""
apply-agrees.py -- produce the submit-ready LQA report.

The Starling Copilot extension stamps "agree" into Column I, but its in-browser
"Export form" download strips all of Excel's formatting (bold headers, colours,
column widths). This script copies those agrees onto the ORIGINAL report -- which
keeps its formatting -- so the file you hand back looks like the LQA report you
received, with the agree column filled in.

USAGE
    python apply-agrees.py  <ORIGINAL.xlsx>  <EXPORT.xlsx>  [OUTPUT.xlsx]

    ORIGINAL  the untouched LQA report you received (keeps its formatting)
    EXPORT    the file the extension's "Export form" button downloaded
    OUTPUT    optional; defaults to "<ORIGINAL name> with-agrees.xlsx"
              next to the original.

Notes
  * It matches rows by Key and never overwrites a Column I cell you already
    filled in by hand.
  * It auto-detects the data tab (e.g. "All") and the Key / "Validation feedback
    (from proofreader)" columns by their headers, so it survives small layout
    changes month to month.
  * It warns if any key in the export isn't found in the original.
"""
import sys
import os
import re

try:
    import openpyxl
except ImportError:
    sys.exit('openpyxl is not installed. Run:  pip install openpyxl')


def s(v):
    return '' if v is None else str(v)


def find_header(rows):
    """Return (header_row_index, header_cells) for the LQA header, else (None, None)."""
    for i, row in enumerate(rows[:10]):
        cells = [s(c) for c in row]
        has_key = any(re.search(r'key', c, re.I) for c in cells)
        has_lqa = any(re.search(r'suggested translation|validation feedback', c, re.I) for c in cells)
        if has_key and has_lqa:
            return i, cells
    return None, None


def col_index(header, pattern):
    for i, h in enumerate(header):
        if re.search(pattern, h, re.I):
            return i
    return -1


def is_lqa_header(header):
    """Strict LQA-report header, like the extension: needs 'Suggested translation'.
    The per-language template tabs use 'Before'/'After' and must NOT qualify, even
    though they carry a 'Validation feedback' column."""
    h = [x.lower() for x in header]
    has = lambda rx: any(re.search(rx, x) for x in h)
    return has('suggested translation') and has('before translation|lqa comment|validation feedback')


def lqa_tab(wb):
    """Name of the LQA-report tab with the most data (e.g. 'All'), else None.
    Ranking by data-row count skips the empty per-language template tabs."""
    best, best_rows = None, -1
    for name in wb.sheetnames:
        rows = list(wb[name].iter_rows(values_only=True, max_row=2000))
        hi, hdr = find_header(rows)
        if hi is None or not is_lqa_header(hdr):
            continue
        data = sum(1 for r in rows[hi + 1:] if any(s(c).strip() for c in r))
        if data > best_rows:
            best, best_rows = name, data
    return best


def main():
    if len(sys.argv) < 3:
        sys.exit(__doc__)

    a_path, b_path = sys.argv[1], sys.argv[2]
    for p in (a_path, b_path):
        if not os.path.exists(p):
            sys.exit('File not found: ' + p)

    # Order-agnostic: the EXPORT is whichever file has more filled Column I cells (the agrees);
    # the ORIGINAL is the other one. Lets a drag-and-drop launcher pass files in any order.
    def col_i_fills(path):
        try:
            wb = openpyxl.load_workbook(path, data_only=True)
            tab = lqa_tab(wb) or wb.sheetnames[0]
            rows = list(wb[tab].iter_rows(values_only=True))
            hi, hdr = find_header(rows)
            if hi is None:
                return 0
            ci = col_index(hdr, r'validation feedback')
            return sum(1 for r in rows[hi + 1:] if 0 <= ci < len(r) and s(r[ci]).strip())
        except Exception:
            return 0
    if col_i_fills(a_path) >= col_i_fills(b_path):
        exp_path, orig_path = a_path, b_path
    else:
        exp_path, orig_path = b_path, a_path

    if len(sys.argv) > 3:
        out_path = sys.argv[3]
    else:
        stem = os.path.splitext(os.path.basename(orig_path))[0]
        out_path = os.path.join(os.path.dirname(orig_path) or '.', stem + ' with-agrees.xlsx')

    # 1) read the agrees the extension stamped (key -> value) from the export.
    # NOT read_only: we scan sheets twice (tab detection, then the read), and a read_only
    # worksheet can only be streamed once -- which silently yields zero rows the second time.
    exp = openpyxl.load_workbook(exp_path, data_only=True)
    etab = lqa_tab(exp) or exp.sheetnames[0]
    erows = list(exp[etab].iter_rows(values_only=True))
    ehi, ehdr = find_header(erows)
    if ehi is None:
        sys.exit('Could not find the LQA header row in the export: ' + exp_path)
    e_key = col_index(ehdr, r'^key$|\bkey\b')
    e_val = col_index(ehdr, r'validation feedback')      # first match = "(from proofreader)" = Col I
    if e_key < 0 or e_val < 0:
        sys.exit('Could not find the Key / Validation-feedback columns in the export.')

    agrees = {}
    for r in erows[ehi + 1:]:
        k = s(r[e_key]).strip() if e_key < len(r) else ''
        v = s(r[e_val]).strip() if e_val < len(r) else ''
        if k and v and k not in agrees:
            agrees[k] = v
    if not agrees:
        sys.exit("No values found in the export's Column I -- nothing to apply.")

    # 2) write them onto the FORMATTED original (openpyxl keeps styling)
    orig = openpyxl.load_workbook(orig_path)
    otab = lqa_tab(orig) or 'All'
    ws = orig[otab]
    ohdr_rows = list(ws.iter_rows(values_only=True, max_row=10))
    ohi, ohdr = find_header(ohdr_rows)
    if ohi is None:
        sys.exit('Could not find the LQA header row in the original: ' + orig_path)
    o_key = col_index(ohdr, r'^key$|\bkey\b') + 1        # openpyxl is 1-based
    o_col = col_index(ohdr, r'validation feedback') + 1
    if o_key < 1 or o_col < 1:
        sys.exit('Could not find the Key / Validation-feedback columns in the original.')

    applied = kept = 0
    used = set()
    for xlrow in range(ohi + 2, ws.max_row + 1):
        k = s(ws.cell(row=xlrow, column=o_key).value).strip()
        if k in agrees:
            cur = s(ws.cell(row=xlrow, column=o_col).value).strip()
            if cur:
                kept += 1                                 # never clobber your own comment
            else:
                ws.cell(row=xlrow, column=o_col).value = agrees[k]
                applied += 1
            used.add(k)
    missing = [k for k in agrees if k not in used]

    orig.save(out_path)

    print('Tab: %r   applied %d agree(s) into Column I   (kept %d existing cell(s) untouched)'
          % (otab, applied, kept))
    if missing:
        print('WARNING: %d key(s) from the export were not found in the original:' % len(missing))
        for k in missing[:10]:
            print('    ' + k)
        if len(missing) > 10:
            print('    ... and %d more' % (len(missing) - 10))
    print('\nSaved: ' + out_path)
    print('This is the file to submit -- original formatting, Column I filled.')


if __name__ == '__main__':
    main()
